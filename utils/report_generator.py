"""
测试报告生成器
支持生成HTML和Excel格式的测试报告
"""

import os
import json
from datetime import datetime
from typing import Dict, List


class ReportGenerator:
    """测试报告生成器"""

    def __init__(
        self,
        stats: Dict,
        global_stats: Dict,
        hostnames: List[str],
        total_cycles: int,
        test_type: str = None,
    ):
        """
        初始化报告生成器

        Args:
            stats: 每台主机的统计信息
            global_stats: 全局统计信息
            hostnames: 主机名列表
            total_cycles: 计划循环次数
            test_type: 测试类型（"s3"或"reboot"），如果不指定则自动判断
        """
        self.stats = stats
        self.global_stats = global_stats
        self.hostnames = hostnames
        self.total_cycles = total_cycles
        # 自动判断测试类型
        if test_type is None:
            # 通过检查调用栈来判断（简单方法：检查stats中是否有s3相关的信息）
            import inspect

            frame = inspect.currentframe()
            try:
                caller = frame.f_back
                caller_name = caller.f_code.co_name if caller else ""
                if "s3" in caller_name.lower() or any(
                    "s3" in str(k).lower() for k in self.stats.keys()
                ):
                    self.test_type = "s3"
                else:
                    self.test_type = "reboot"
            except:
                self.test_type = "reboot"  # 默认
        else:
            self.test_type = test_type.lower()
        self.report_dir = "reports"
        if not os.path.exists(self.report_dir):
            os.makedirs(self.report_dir)

    def _calculate_statistics(self, hostname: str) -> Dict:
        """计算统计信息"""
        cycle_durations = self.stats[hostname].get("cycle_durations", [])
        if not cycle_durations:
            return {
                "avg_duration": 0,
                "min_duration": 0,
                "max_duration": 0,
                "total_duration": 0,
                "success_count": 0,
                "failed_count": 0,
            }

        durations = [d["duration"] for d in cycle_durations]
        success_durations = [
            d["duration"] for d in cycle_durations if d["status"] == "success"
        ]
        failed_durations = [
            d["duration"] for d in cycle_durations if d["status"] == "failed"
        ]

        return {
            "avg_duration": sum(durations) / len(durations) if durations else 0,
            "min_duration": min(durations) if durations else 0,
            "max_duration": max(durations) if durations else 0,
            "total_duration": sum(durations),
            "success_avg": (
                sum(success_durations) / len(success_durations)
                if success_durations
                else 0
            ),
            "failed_avg": (
                sum(failed_durations) / len(failed_durations) if failed_durations else 0
            ),
            "success_count": len(success_durations),
            "failed_count": len(failed_durations),
        }

    def generate_html_report(self):
        """生成HTML报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_file = os.path.join(
            self.report_dir, f"{self.test_type}_test_report_{timestamp}.html"
        )

        # 计算汇总统计
        total_success = sum(s["success"] for s in self.stats.values())
        total_failed = sum(s["failed"] for s in self.stats.values())
        total_critical = sum(s["critical_failures"] for s in self.stats.values())
        total_recoverable = sum(
            s["non_critical_recoverable_failures"] for s in self.stats.values()
        )
        total_unrecoverable = sum(
            s["non_critical_unrecoverable_failures"] for s in self.stats.values()
        )

        # 确定测试类型显示名称
        if self.test_type == "s3":
            test_type_display = "S3"
        elif self.test_type == "s4":
            test_type_display = "S4"
        else:
            test_type_display = "重启"

        # 生成HTML内容
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{test_type_display}测试报告</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4CAF50;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .success {{
            color: #4CAF50;
            font-weight: bold;
        }}
        .failed {{
            color: #f44336;
            font-weight: bold;
        }}
        .warning {{
            color: #ff9800;
            font-weight: bold;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        .summary-card {{
            background-color: #f9f9f9;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #4CAF50;
        }}
        .summary-card h3 {{
            margin: 0 0 10px 0;
            color: #555;
        }}
        .summary-card .value {{
            font-size: 24px;
            font-weight: bold;
            color: #333;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{test_type_display}测试报告</h1>
        
        <div class="summary">
            <div class="summary-card">
                <h3>测试时间</h3>
                <div class="value">{self.global_stats.get('start_time', 'N/A')}</div>
                <div>至 {self.global_stats.get('end_time', 'N/A')}</div>
            </div>
            <div class="summary-card">
                <h3>测试主机</h3>
                <div class="value">{len(self.hostnames)}</div>
                <div>{', '.join(self.hostnames)}</div>
            </div>
            <div class="summary-card">
                <h3>计划循环次数</h3>
                <div class="value">{self.total_cycles}</div>
            </div>
            <div class="summary-card">
                <h3>成功次数</h3>
                <div class="value success">{total_success}</div>
            </div>
            <div class="summary-card">
                <h3>失败次数</h3>
                <div class="value failed">{total_failed}</div>
            </div>
        </div>

        <h2>汇总统计</h2>
        <table>
            <tr>
                <th>项目</th>
                <th>数量</th>
            </tr>
            <tr>
                <td>成功次数</td>
                <td class="success">{total_success}</td>
            </tr>
            <tr>
                <td>失败次数</td>
                <td class="failed">{total_failed}</td>
            </tr>
            <tr>
                <td>Critical失败</td>
                <td class="failed">{total_critical}</td>
            </tr>
            <tr>
                <td>非Critical可恢复失败</td>
                <td class="warning">{total_recoverable}</td>
            </tr>
            <tr>
                <td>非Critical不可恢复失败</td>
                <td class="failed">{total_unrecoverable}</td>
            </tr>
        </table>
"""

        # 为每台主机生成详细报告
        for hostname in self.hostnames:
            stats = self.stats[hostname]
            calc_stats = self._calculate_statistics(hostname)
            cycle_durations = stats.get("cycle_durations", [])

            html_content += f"""
        <h2>{hostname} 详细统计</h2>
        <table>
            <tr>
                <th>项目</th>
                <th>值</th>
            </tr>
            <tr>
                <td>完成次数</td>
                <td>{stats['total']}</td>
            </tr>
            <tr>
                <td>成功次数</td>
                <td class="success">{stats['success']}</td>
            </tr>
            <tr>
                <td>失败次数</td>
                <td class="failed">{stats['failed']}</td>
            </tr>
            <tr>
                <td>平均耗时（秒）</td>
                <td>{calc_stats['avg_duration']:.2f}</td>
            </tr>
            <tr>
                <td>最小耗时（秒）</td>
                <td>{calc_stats['min_duration']:.2f}</td>
            </tr>
            <tr>
                <td>最大耗时（秒）</td>
                <td>{calc_stats['max_duration']:.2f}</td>
            </tr>
            <tr>
                <td>成功平均耗时（秒）</td>
                <td>{calc_stats['success_avg']:.2f}</td>
            </tr>
            <tr>
                <td>失败平均耗时（秒）</td>
                <td>{calc_stats['failed_avg']:.2f}</td>
            </tr>
        </table>

        <h3>{hostname} 每次{test_type_display}耗时详情</h3>
        <table>
            <tr>
                <th>循环次数</th>
                <th>状态</th>
                <th>耗时（秒）</th>
                <th>时间戳</th>
                <th>备注</th>
            </tr>
"""

            for cycle_data in cycle_durations:
                status_class = (
                    "success" if cycle_data["status"] == "success" else "failed"
                )
                status_text = "成功" if cycle_data["status"] == "success" else "失败"
                error_info = cycle_data.get("error", "")
                failure_type = cycle_data.get("failure_type", "")
                remark = f"{failure_type}: {error_info}" if error_info else ""

                html_content += f"""
            <tr>
                <td>{cycle_data['cycle']}</td>
                <td class="{status_class}">{status_text}</td>
                <td>{cycle_data['duration']:.2f}</td>
                <td>{cycle_data['timestamp']}</td>
                <td>{remark}</td>
            </tr>
"""

            html_content += """
        </table>
"""

        html_content += """
    </div>
</body>
</html>
"""

        with open(html_file, "w", encoding="utf-8") as f:
            f.write(html_content)

        print(f"HTML报告已生成: {html_file}")

    def generate_excel_report(self):
        """生成Excel报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("未安装openpyxl，无法生成Excel报告。请运行: pip install openpyxl")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(
            self.report_dir, f"{self.test_type}_test_report_{timestamp}.xlsx"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总统计"

        # 设置标题样式
        header_fill = PatternFill(
            start_color="4CAF50", end_color="4CAF50", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # 确定测试类型显示名称
        if self.test_type == "s3":
            test_type_display = "S3"
        elif self.test_type == "s4":
            test_type_display = "S4"
        else:
            test_type_display = "重启"

        # 汇总统计表
        ws.append([f"{test_type_display}测试报告"])
        ws.merge_cells("A1:E1")
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([])
        ws.append(["测试时间", str(self.global_stats.get("start_time", "N/A"))])
        ws.append(["结束时间", str(self.global_stats.get("end_time", "N/A"))])
        ws.append(["测试主机", ", ".join(self.hostnames)])
        ws.append(["计划循环次数", self.total_cycles])
        ws.append([])

        # 汇总数据
        total_success = sum(s["success"] for s in self.stats.values())
        total_failed = sum(s["failed"] for s in self.stats.values())
        total_critical = sum(s["critical_failures"] for s in self.stats.values())
        total_recoverable = sum(
            s["non_critical_recoverable_failures"] for s in self.stats.values()
        )
        total_unrecoverable = sum(
            s["non_critical_unrecoverable_failures"] for s in self.stats.values()
        )

        ws.append(["项目", "数量"])
        ws.append(["成功次数", total_success])
        ws.append(["失败次数", total_failed])
        ws.append(["Critical失败", total_critical])
        ws.append(["非Critical可恢复失败", total_recoverable])
        ws.append(["非Critical不可恢复失败", total_unrecoverable])

        # 为每台主机创建详细工作表
        for hostname in self.hostnames:
            stats = self.stats[hostname]
            calc_stats = self._calculate_statistics(hostname)
            cycle_durations = stats.get("cycle_durations", [])

            ws_detail = wb.create_sheet(title=hostname)

            # 统计信息
            ws_detail.append([f"{hostname} 详细统计"])
            ws_detail.merge_cells("A1:B1")
            ws_detail["A1"].font = Font(size=14, bold=True)
            ws_detail.append([])
            ws_detail.append(["项目", "值"])
            ws_detail.append(["完成次数", stats["total"]])
            ws_detail.append(["成功次数", stats["success"]])
            ws_detail.append(["失败次数", stats["failed"]])
            ws_detail.append(["平均耗时（秒）", f"{calc_stats['avg_duration']:.2f}"])
            ws_detail.append(["最小耗时（秒）", f"{calc_stats['min_duration']:.2f}"])
            ws_detail.append(["最大耗时（秒）", f"{calc_stats['max_duration']:.2f}"])
            ws_detail.append(["成功平均耗时（秒）", f"{calc_stats['success_avg']:.2f}"])
            ws_detail.append(["失败平均耗时（秒）", f"{calc_stats['failed_avg']:.2f}"])
            ws_detail.append([])

            # 每次操作耗时详情
            ws_detail.append(["循环次数", "状态", "耗时（秒）", "时间戳", "备注"])
            header_row = ws_detail.max_row
            for col in range(1, 6):
                cell = ws_detail.cell(row=header_row, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for cycle_data in cycle_durations:
                status_text = "成功" if cycle_data["status"] == "success" else "失败"
                error_info = cycle_data.get("error", "")
                failure_type = cycle_data.get("failure_type", "")
                remark = f"{failure_type}: {error_info}" if error_info else ""

                ws_detail.append(
                    [
                        cycle_data["cycle"],
                        status_text,
                        f"{cycle_data['duration']:.2f}",
                        str(cycle_data["timestamp"]),
                        remark,
                    ]
                )

            # 设置列宽
            ws_detail.column_dimensions["A"].width = 12
            ws_detail.column_dimensions["B"].width = 10
            ws_detail.column_dimensions["C"].width = 15
            ws_detail.column_dimensions["D"].width = 25
            ws_detail.column_dimensions["E"].width = 50

        # 设置汇总表的列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        wb.save(excel_file)
        print(f"Excel报告已生成: {excel_file}")
